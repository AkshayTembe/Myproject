#!/usr/bin/env python3
"""
DBC to Excel Converter
Converts a DBC file into Excel format compatible with ExcelParser C# library.

Usage:
    python dbc_to_excel.py input.dbc output.xlsx

Output Excel structure matches ExcelParser sheet parsers:
- Nodes
- ValueTables
- Messages
- Signals
- ExtraTransmitters
- EnvironmentVariables
- BA_DEF
- BA
- Comments
"""

import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict


class DbcParser:
    """Simple DBC file parser"""
    
    def __init__(self, dbc_path):
        self.dbc_path = dbc_path
        self.nodes = []
        self.messages = []
        self.signals = {}  # msg_id -> [signals]
        self.value_tables = {}  # name -> {int: str}
        self.signal_value_tables = {}  # (msg_id, signal_name) -> {int: str}
        self.env_var_value_tables = {}  # env_name -> {int: str}
        self.comments = []  # (type, scope, comment)
        self.ba_defs = []  # (scope, name, type, min, max, enum_vals)
        self.ba_assignments = []  # (scope, scope_id, attr_name, value)
        self.extra_transmitters = {}  # msg_id -> [transmitters]
        self.env_vars = []  # environment variables
        self.signal_value_types = {}  # (msg_id, signal_name) -> value_type
        
    def parse(self):
        """Parse DBC file"""
        with open(self.dbc_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        self._parse_nodes(content)
        self._parse_value_tables(content)
        self._parse_messages_and_signals(content)
        self._parse_signal_val_entries(content)
        self._parse_env_var_val_entries(content)
        self._parse_comments(content)
        self._parse_ba_defs(content)
        self._parse_ba_assignments(content)
        self._parse_extra_transmitters(content)
        self._parse_env_vars(content)
        self._parse_signal_value_types(content)
        
    def _parse_nodes(self, content):
        """Parse BU_ line"""
        match = re.search(r'BU_:\s*(.+)', content)
        if match:
            nodes_str = match.group(1).strip()
            if nodes_str:
                self.nodes = [n.strip() for n in nodes_str.split() if n.strip()]
    
    def _parse_value_tables(self, content):
        """Parse VAL_TABLE_ entries"""
        pattern = r'VAL_TABLE_\s+(\S+)\s+(.+?)\s*;'
        for match in re.finditer(pattern, content, re.MULTILINE):
            table_name = match.group(1)
            values_str = match.group(2)
            self.value_tables[table_name] = self._parse_value_dict(values_str)
    
    def _parse_value_dict(self, values_str):
        """Parse value dictionary from string like: 0 "Off" 1 "On" """
        result = {}
        pattern = r'(-?\d+)\s+"([^"]*)"'
        for match in re.finditer(pattern, values_str):
            key = int(match.group(1))
            value = match.group(2)
            result[key] = value
        return result
    
    def _parse_messages_and_signals(self, content):
        """Parse BO_ messages and SG_ signals"""
        # Parse messages
        msg_pattern = r'BO_\s+(\d+)\s+(\S+):\s*(\d+)\s+(\S+)'
        for match in re.finditer(msg_pattern, content, re.MULTILINE):
            msg_id_raw = int(match.group(1))
            # Handle extended ID flag
            is_extended = msg_id_raw >= 0x80000000
            msg_id = msg_id_raw - 0x80000000 if is_extended else msg_id_raw
            
            msg = {
                'id': msg_id,
                'id_hex': f"0x{msg_id:X}",
                'name': match.group(2),
                'dlc': int(match.group(3)),
                'transmitter': match.group(4),
                'is_extended': is_extended,
                'comment': ''
            }
            self.messages.append(msg)
            self.signals[msg_id] = []
        
        # Parse signals
        sig_pattern = r'SG_\s+(\S+)\s*(M|m\d+M?)?\s*:\s*(\d+)\|(\d+)@([01])([+-])\s*\(([^,]+),([^)]+)\)\s*\[([^|]+)\|([^\]]+)\]\s*"([^"]*)"\s*(.+)'
        
        for match in re.finditer(sig_pattern, content, re.MULTILINE):
            sig_name = match.group(1)
            mux = match.group(2) or ''
            start_bit = int(match.group(3))
            length = int(match.group(4))
            byte_order = match.group(5)
            sign = match.group(6)
            factor = match.group(7).strip()
            offset = match.group(8).strip()
            minimum = match.group(9).strip()
            maximum = match.group(10).strip()
            unit = match.group(11)
            receivers_str = match.group(12).strip()
            
            # Find which message this belongs to (search backwards)
            msg_id = None
            sig_pos = match.start()
            for msg_match in re.finditer(msg_pattern, content[:sig_pos], re.MULTILINE):
                msg_id_raw = int(msg_match.group(1))
                msg_id = msg_id_raw - 0x80000000 if msg_id_raw >= 0x80000000 else msg_id_raw
            
            if msg_id is not None and msg_id in self.signals:
                receivers = [r.strip() for r in receivers_str.replace(',', ' ').split() if r.strip()]
                
                signal = {
                    'msg_id': msg_id,
                    'name': sig_name,
                    'start_bit': start_bit,
                    'length': length,
                    'byte_order_sign': f"@{byte_order}{sign}",
                    'factor_offset': f"({factor},{offset})",
                    'min_max': f"[{minimum}|{maximum}]",
                    'unit': unit,
                    'receivers': ','.join(receivers),
                    'multiplexing': mux,
                    'comment': '',
                    'value_table': ''
                }
                self.signals[msg_id].append(signal)
    
    def _parse_signal_val_entries(self, content):
        """Parse VAL_ entries for signals"""
        pattern = r'VAL_\s+(\d+)\s+(\S+)\s+(.+?)\s*;'
        for match in re.finditer(pattern, content, re.MULTILINE):
            msg_id_raw = int(match.group(1))
            msg_id = msg_id_raw - 0x80000000 if msg_id_raw >= 0x80000000 else msg_id_raw
            sig_name = match.group(2)
            values_str = match.group(3)
            self.signal_value_tables[(msg_id, sig_name)] = self._parse_value_dict(values_str)
    
    def _parse_env_var_val_entries(self, content):
        """Parse VAL_ entries for environment variables"""
        pattern = r'VAL_\s+(\S+)\s+(.+?)\s*;'
        for match in re.finditer(pattern, content, re.MULTILINE):
            first_token = match.group(1)
            # Check if it's a number (message ID) or string (env var)
            if not first_token.isdigit():
                env_name = first_token
                values_str = match.group(2)
                self.env_var_value_tables[env_name] = self._parse_value_dict(values_str)
    
    def _parse_comments(self, content):
        """Parse CM_ entries"""
        # Message comments
        pattern = r'CM_\s+BO_\s+(\d+)\s+"([^"]*)"\s*;'
        for match in re.finditer(pattern, content, re.MULTILINE):
            msg_id_raw = int(match.group(1))
            msg_id = msg_id_raw - 0x80000000 if msg_id_raw >= 0x80000000 else msg_id_raw
            comment = match.group(2)
            self.comments.append(('BO', str(msg_id), comment))
        
        # Signal comments
        pattern = r'CM_\s+SG_\s+(\d+)\s+(\S+)\s+"([^"]*)"\s*;'
        for match in re.finditer(pattern, content, re.MULTILINE):
            msg_id_raw = int(match.group(1))
            msg_id = msg_id_raw - 0x80000000 if msg_id_raw >= 0x80000000 else msg_id_raw
            sig_name = match.group(2)
            comment = match.group(3)
            self.comments.append(('SG', f"{msg_id}:{sig_name}", comment))
        
        # Node comments
        pattern = r'CM_\s+BU_\s+(\S+)\s+"([^"]*)"\s*;'
        for match in re.finditer(pattern, content, re.MULTILINE):
            node = match.group(1)
            comment = match.group(2)
            self.comments.append(('BU', node, comment))
        
        # Env var comments
        pattern = r'CM_\s+EV_\s+(\S+)\s+"([^"]*)"\s*;'
        for match in re.finditer(pattern, content, re.MULTILINE):
            env = match.group(1)
            comment = match.group(2)
            self.comments.append(('EV', env, comment))
    
    def _parse_ba_defs(self, content):
        """Parse BA_DEF_ entries"""
        # BA_DEF_ [scope] "name" type [params] ;
        pattern = r'BA_DEF_\s+(BU_|BO_|SG_|EV_)?\s*"([^"]+)"\s+(\S+)\s*([^;]*);'
        for match in re.finditer(pattern, content, re.MULTILINE):
            scope_raw = match.group(1) or ''
            scope = scope_raw.strip('_').upper() if scope_raw else 'GLOBAL'
            if scope == 'BO': scope = 'MESSAGE'
            elif scope == 'BU': scope = 'NODE'
            elif scope == 'SG': scope = 'SIGNAL'
            elif scope == 'EV': scope = 'ENV'
            
            name = match.group(2)
            type_str = match.group(3).upper()
            params = match.group(4).strip()
            
            min_val = max_val = enum_vals = ''
            
            if type_str in ['INT', 'HEX', 'FLOAT']:
                parts = params.split()
                if len(parts) >= 2:
                    min_val = parts[0]
                    max_val = parts[1]
            elif type_str == 'ENUM':
                # Extract enum values
                enum_vals = params
            
            self.ba_defs.append((scope, name, type_str, min_val, max_val, enum_vals))
    
    def _parse_ba_assignments(self, content):
        """Parse BA_ entries"""
        # BA_ "name" [scope] [scope_id] value ;
        
        # Global
        pattern = r'BA_\s+"([^"]+)"\s+([^;]+);'
        for match in re.finditer(pattern, content, re.MULTILINE):
            attr_name = match.group(1)
            rest = match.group(2).strip()
            
            # Check if it's a scoped assignment
            if rest.startswith('BU_'):
                scope = 'BU'
                parts = rest.split(None, 2)
                scope_id = parts[1] if len(parts) > 1 else ''
                value = parts[2] if len(parts) > 2 else ''
            elif rest.startswith('BO_'):
                scope = 'BO'
                parts = rest.split(None, 2)
                scope_id = parts[1] if len(parts) > 1 else ''
                value = parts[2] if len(parts) > 2 else ''
            elif rest.startswith('SG_'):
                scope = 'SG'
                parts = rest.split(None, 3)
                msg_id = parts[1] if len(parts) > 1 else ''
                sig_name = parts[2] if len(parts) > 2 else ''
                msg_id_int = int(msg_id)
                if msg_id_int >= 0x80000000:
                    msg_id_int -= 0x80000000
                scope_id = f"{msg_id_int}:{sig_name}"
                value = parts[3] if len(parts) > 3 else ''
            elif rest.startswith('EV_'):
                scope = 'EV'
                parts = rest.split(None, 2)
                scope_id = parts[1] if len(parts) > 1 else ''
                value = parts[2] if len(parts) > 2 else ''
            else:
                # Global assignment
                scope = 'GLOBAL'
                scope_id = ''
                value = rest
            
            # Clean value
            value = value.strip().strip('"')
            self.ba_assignments.append((scope, scope_id, attr_name, value))
    
    def _parse_extra_transmitters(self, content):
        """Parse BO_TX_BU_ entries"""
        pattern = r'BO_TX_BU_\s+(\d+)\s*:\s*([^;]+);'
        for match in re.finditer(pattern, content, re.MULTILINE):
            msg_id_raw = int(match.group(1))
            msg_id = msg_id_raw - 0x80000000 if msg_id_raw >= 0x80000000 else msg_id_raw
            transmitters_str = match.group(2).strip()
            transmitters = [t.strip() for t in transmitters_str.replace(',', ' ').split() if t.strip()]
            self.extra_transmitters[msg_id] = transmitters
    
    def _parse_env_vars(self, content):
        """Parse EV_ entries"""
        pattern = r'EV_\s+(\S+)\s*:\s*(\d+)\s*\[([^|]+)\|([^\]]+)\]\s*"([^"]*)"\s+(\S+)\s+(\d+)\s+(\S+)\s+(.+?);'
        for match in re.finditer(pattern, content, re.MULTILINE):
            env = {
                'name': match.group(1),
                'type': match.group(2),  # 0=int, 1=float, 2=string
                'min': match.group(3).strip(),
                'max': match.group(4).strip(),
                'unit': match.group(5),
                'default': match.group(6),
                'access': match.group(7),
                'nodes': match.group(9).strip() if len(match.groups()) > 8 else '',
                'comment': ''
            }
            self.env_vars.append(env)
    
    def _parse_signal_value_types(self, content):
        """Parse SIG_VALTYPE_ entries"""
        pattern = r'SIG_VALTYPE_\s+(\d+)\s+(\S+)\s*:\s*(\d+)\s*;'
        for match in re.finditer(pattern, content, re.MULTILINE):
            msg_id_raw = int(match.group(1))
            msg_id = msg_id_raw - 0x80000000 if msg_id_raw >= 0x80000000 else msg_id_raw
            sig_name = match.group(2)
            val_type = int(match.group(3))  # 1=float, 2=double
            self.signal_value_types[(msg_id, sig_name)] = val_type


def create_excel_from_dbc(dbc_path, output_path):
    """Convert DBC file to Excel format compatible with ExcelParser"""
    
    print(f"Parsing DBC file: {dbc_path}")
    parser = DbcParser(dbc_path)
    parser.parse()
    
    print(f"Creating Excel workbook: {output_path}")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 1. Nodes sheet
    print("  Writing Nodes sheet...")
    ws = wb.create_sheet("Nodes")
    ws.append(["NodeName"])
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    for node in parser.nodes:
        ws.append([node])
    
    # 2. ValueTables sheet
    print("  Writing ValueTables sheet...")
    ws = wb.create_sheet("ValueTables")
    ws.append(["TableName", "Values (format: key:\"value\" key:\"value\")"])
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill
    
    for table_name, values in parser.value_tables.items():
        values_str = ' '.join([f'{k}:"{v}"' for k, v in sorted(values.items())])
        ws.append([table_name, values_str])
    
    # 3. Messages sheet
    print("  Writing Messages sheet...")
    ws = wb.create_sheet("Messages")
    ws.append(["MessageID", "MessageName", "DLC", "Transmitter", "IsExtended", "Comment"])
    for col in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
        ws[col].font = header_font
        ws[col].fill = header_fill
    
    for msg in parser.messages:
        ws.append([
            msg['id_hex'],
            msg['name'],
            msg['dlc'],
            msg['transmitter'],
            'TRUE' if msg['is_extended'] else 'FALSE',
            ''  # Comment filled later
        ])
    
    # 4. Signals sheet
    print("  Writing Signals sheet...")
    ws = wb.create_sheet("Signals")
    headers = ["MessageID", "SignalName", "StartBit", "Length", "ByteOrder@Sign", 
               "Factor,Offset", "Min|Max", "Unit", "Receivers", "Comment", 
               "InitialValue", "ValueType", "SendType", "Multiplexing", "ValueTable"]
    ws.append(headers)
    for i, col in enumerate(['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1'], 1):
        ws[col].font = header_font
        ws[col].fill = header_fill
    
    for msg_id in sorted(parser.signals.keys()):
        for sig in parser.signals[msg_id]:
            # Check if there's a value table for this signal
            value_table_name = ''
            if (msg_id, sig['name']) in parser.signal_value_tables:
                # Create inline value table or reference
                vtable = parser.signal_value_tables[(msg_id, sig['name'])]
                value_table_name = ' '.join([f'{k}:"{v}"' for k, v in sorted(vtable.items())])
            
            ws.append([
                f"0x{msg_id:X}",
                sig['name'],
                sig['start_bit'],
                sig['length'],
                sig['byte_order_sign'],
                sig['factor_offset'],
                sig['min_max'],
                sig['unit'],
                sig['receivers'],
                '',  # Comment filled later
                '',  # InitialValue
                '',  # ValueType
                '',  # SendType
                sig['multiplexing'],
                value_table_name
            ])
    
    # 5. ExtraTransmitters sheet
    if parser.extra_transmitters:
        print("  Writing ExtraTransmitters sheet...")
        ws = wb.create_sheet("ExtraTransmitters")
        ws.append(["MessageID", "AdditionalTransmitters"])
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].fill = header_fill
        
        for msg_id, transmitters in parser.extra_transmitters.items():
            ws.append([f"0x{msg_id:X}", ','.join(transmitters)])
    
    # 6. EnvironmentVariables sheet
    if parser.env_vars:
        print("  Writing EnvironmentVariables sheet...")
        ws = wb.create_sheet("EnvironmentVariables")
        ws.append(["Name", "Type", "Min", "Max", "Default", "Unit", "Nodes", "DataLength", "Comment"])
        for col in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1']:
            ws[col].font = header_font
            ws[col].fill = header_fill
        
        for env in parser.env_vars:
            env_type = {'0': 'INT', '1': 'FLOAT', '2': 'STRING'}.get(env['type'], 'INT')
            ws.append([
                env['name'],
                env_type,
                env['min'],
                env['max'],
                env['default'],
                env['unit'],
                env['nodes'],
                '',  # DataLength
                ''   # Comment
            ])
    
    # 7. BA_DEF sheet
    if parser.ba_defs:
        print("  Writing BA_DEF sheet...")
        ws = wb.create_sheet("BA_DEF")
        ws.append(["Scope", "PropertyName", "Type", "Min", "Max", "EnumValues"])
        for col in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
            ws[col].font = header_font
            ws[col].fill = header_fill
        
        for scope, name, type_str, min_val, max_val, enum_vals in parser.ba_defs:
            ws.append([scope, name, type_str, min_val, max_val, enum_vals])
    
    # 8. BA sheet (assignments)
    if parser.ba_assignments:
        print("  Writing BA sheet...")
        ws = wb.create_sheet("BA")
        ws.append(["Scope", "ScopeIdentifier", "AttributeName", "Value"])
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws[col].font = header_font
            ws[col].fill = header_fill
        
        for scope, scope_id, attr_name, value in parser.ba_assignments:
            ws.append([scope, scope_id, attr_name, value])
    
    # 9. Comments sheet
    if parser.comments:
        print("  Writing Comments sheet...")
        ws = wb.create_sheet("Comments")
        ws.append(["Type", "Scope", "Comment"])
        for col in ['A1', 'B1', 'C1']:
            ws[col].font = header_font
            ws[col].fill = header_fill
        
        for comment_type, scope, comment in parser.comments:
            ws.append([comment_type, scope, comment])
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel file created successfully: {output_path}")
    print(f"  Nodes: {len(parser.nodes)}")
    print(f"  Messages: {len(parser.messages)}")
    print(f"  Signals: {sum(len(sigs) for sigs in parser.signals.values())}")
    print(f"  Value Tables: {len(parser.value_tables)}")
    print(f"  Custom Properties: {len(parser.ba_defs)}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python dbc_to_excel.py input.dbc output.xlsx")
        sys.exit(1)
    
    input_file = Path(sys.argv[1])
    output_file = Path(sys.argv[2])
    
    if not input_file.exists():
        print(f"Error: Input file not found: {input_file}")
        sys.exit(1)
    
    try:
        create_excel_from_dbc(str(input_file), str(output_file))
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)